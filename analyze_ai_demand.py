"""
Analysis of AI skills demand in Polish job market based on ESCO skills + keyword matching.
"""

import sqlite3
import json
import re
import pandas as pd
import numpy as np
from collections import Counter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

print("Loading data...")

# ── 1. Load AI skills reference ──────────────────────────────────────────────
ai_ref = pd.read_csv("esco_ai_skills.csv", sep=";")
pl_skills = pd.read_csv(
    "ESCO dataset - v1.2.1 - classification - pl - csv/skills_pl.csv"
)
ai_ref = ai_ref.merge(pl_skills[["conceptUri", "preferredLabel"]], on="conceptUri", how="left")

ai_skills_pl = {}
for _, row in ai_ref.iterrows():
    label = row["preferredLabel"].strip().lower()
    ai_skills_pl[label] = {
        "name_en": row["ESCO Skill/Knowledge Name"],
        "category": row["Revised Category"],
        "scope": row["Portfolio Scope"],
    }

# Load transversal skills
transversal = pd.read_csv(
    "ESCO dataset - v1.2.1 - classification - pl - csv/transversalSkillsCollection_pl.csv"
)
transversal_set = set(transversal["preferredLabel"].str.strip().str.lower())

# EN labels lookup: skills PL->EN
en_skills = pd.read_csv(
    "ESCO dataset - v1.2.1 - classification - en - csv/skills_en.csv"
)
skill_uri_to_en = dict(zip(en_skills["conceptUri"], en_skills["preferredLabel"]))
skill_pl_to_en = {}
for _, r in pl_skills.iterrows():
    pl_label = str(r["preferredLabel"]).strip().lower()
    uri = r["conceptUri"]
    if uri in skill_uri_to_en:
        skill_pl_to_en[pl_label] = skill_uri_to_en[uri]

# EN labels lookup: occupations PL->EN
occ_pl = pd.read_csv(
    "ESCO dataset - v1.2.1 - classification - pl - csv/occupations_pl.csv"
)
occ_en = pd.read_csv(
    "ESCO dataset - v1.2.1 - classification - en - csv/occupations_en.csv"
)
occ_uri_to_en = dict(zip(occ_en["conceptUri"], occ_en["preferredLabel"]))
occ_pl_to_en = {}
for _, r in occ_pl.iterrows():
    pl_label = str(r["preferredLabel"]).strip().lower()
    uri = r["conceptUri"]
    if uri in occ_uri_to_en:
        occ_pl_to_en[pl_label] = occ_uri_to_en[uri]

# ── 2. Load job ads ──────────────────────────────────────────────────────────
conn = sqlite3.connect("jobs_database.db")
df = pd.read_sql_query(
    """SELECT id, title, requirements, responsibilities, 
              contract_types, seniority_level, location, technologies,
              skills_esco_contextual, title_clean
       FROM job_ads""",
    conn,
)
conn.close()

print(f"Loaded {len(df)} job ads")

# ── 3. Load NACE data ────────────────────────────────────────────────────────
conn_nace = sqlite3.connect("job_nace.db")
nace_df = pd.read_sql_query(
    "SELECT job_id, esco_label, nace_code, nace_title FROM job_nace", conn_nace
)
conn_nace.close()

# ── 4. Parse ESCO skills and flag AI offers ──────────────────────────────────
print("Parsing ESCO skills...")


def parse_skills(raw):
    if not raw or pd.isna(raw):
        return []
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return []


df["skills_parsed"] = df["skills_esco_contextual"].apply(parse_skills)


def check_ai_esco(skills_list):
    found_skills = []
    for s in skills_list:
        esco_label = s.get("esco", "").strip().lower()
        if esco_label in ai_skills_pl:
            info = ai_skills_pl[esco_label]
            found_skills.append(
                {
                    "esco_pl": esco_label,
                    "esco_en": info["name_en"],
                    "category": info["category"],
                    "scope": info["scope"],
                }
            )
    return found_skills


df["ai_skills_esco"] = df["skills_parsed"].apply(check_ai_esco)
df["has_ai_esco"] = df["ai_skills_esco"].apply(lambda x: len(x) > 0)
df["has_ai_core_esco"] = df["ai_skills_esco"].apply(
    lambda x: any(s["scope"] == "Strict AI" for s in x)
)
df["has_ai_extended_esco"] = df["ai_skills_esco"].apply(
    lambda x: any(s["scope"] == "Extended AI" for s in x)
)

# ── 5. Keyword-based AI matching ─────────────────────────────────────────────
print("Keyword matching...")

AI_KEYWORDS_STRICT = [
    r"\bartificial intelligence\b",
    r"\bsztuczna inteligencja\b",
    r"\bmachine learning\b",
    r"\buczenie maszynowe\b",
    r"\bdeep learning\b",
    r"\buczenie głębokie\b",
    r"\bneural network\b",
    r"\bsieć neuronow\w*\b",
    r"\bcomputer vision\b",
    r"\bnatural language processing\b",
    r"\bNLP\b",
    r"\bLLM\b",
    r"\blarge language model\b",
    r"\bGPT\b",
    r"\bChatGPT\b",
    r"\bgenerative ai\b",
    r"\bgen ai\b",
    r"\bMLOps\b",
    r"\bml engineer\b",
    r"\bai engineer\b",
    r"\bdata scien\w+\b",
    r"\btensorflow\b",
    r"\bpytorch\b",
    r"\bscikit-learn\b",
    r"\bkeras\b",
    r"\breinforcement learning\b",
    r"\btransfer learning\b",
    r"\bspeech recognition\b",
    r"\brozpoznawanie mowy\b",
    r"\bimage recognition\b",
    r"\bpredictive model\w*\b",
    r"\bmodel\w* predykcyjn\w*\b",
    r"\bbig data\b",
    r"\bdata mining\b",
    r"\bexploracja danych\b",
    r"\brecommender system\b",
    r"\brecommendation system\b",
    r"\bcognitive computing\b",
    r"\bautonomous vehicle\b",
    r"\bself-driving\b",
    r"\brobotic process automation\b",
]

AI_KEYWORDS_EXTENDED = AI_KEYWORDS_STRICT + [
    r"\bRPA\b",
    r"\bpredictive analytics\b",
    r"\banalityka predykcyjna\b",
    r"\bbiometr\w+\b",
    r"\brobot\w*\b",
]

strict_pattern = re.compile("|".join(AI_KEYWORDS_STRICT), re.IGNORECASE)
extended_pattern = re.compile("|".join(AI_KEYWORDS_EXTENDED), re.IGNORECASE)


def build_search_text(row):
    parts = []
    for col in ["title", "requirements", "responsibilities"]:
        v = row[col]
        if v and not pd.isna(v):
            parts.append(str(v))
    return " ".join(parts)


df["search_text"] = df.apply(build_search_text, axis=1)
df["has_ai_kw_strict"] = df["search_text"].apply(lambda t: bool(strict_pattern.search(t)))
df["has_ai_kw_extended"] = df["search_text"].apply(lambda t: bool(extended_pattern.search(t)))


def find_matched_keywords(text):
    if not text:
        return []
    return list(set(m.group() for m in strict_pattern.finditer(text)))


df["ai_keywords_found"] = df["search_text"].apply(find_matched_keywords)

# Combined flag
df["is_ai_offer"] = df["has_ai_esco"] | df["has_ai_kw_strict"]
df["is_ai_strict"] = df["has_ai_core_esco"] | df["has_ai_kw_strict"]
df["is_ai_extended_only"] = df["is_ai_offer"] & ~df["is_ai_strict"]

# ICT flag: non-null technologies column
df["is_ict"] = df["technologies"].notna() & (df["technologies"].str.strip() != "")
df["is_ai_and_ict"] = df["is_ai_offer"] & df["is_ict"]

print(f"AI offers (ESCO): {df['has_ai_esco'].sum()}")
print(f"AI offers (keywords strict): {df['has_ai_kw_strict'].sum()}")
print(f"AI offers (combined): {df['is_ai_offer'].sum()}")
print(f"ICT offers (has technologies): {df['is_ict'].sum()}")
print(f"AI & ICT overlap: {df['is_ai_and_ict'].sum()}")

# ── 6. ANALYSIS ──────────────────────────────────────────────────────────────
total_offers = len(df)
ai_offers = df[df["is_ai_offer"]]
non_ai = df[~df["is_ai_offer"]]

# ── 6.1 Overview ─────────────────────────────────────────────────────────────
print("\n=== Building analyses ===")

n_ict = int(df["is_ict"].sum())
n_ai_ict = int(df["is_ai_and_ict"].sum())

overview_data = {
    "Metric": [
        "Total job offers",
        "ICT offers (has technologies)",
        "% ICT offers",
        "",
        "AI offers (ESCO match)",
        "AI offers (keyword match)",
        "AI offers (ESCO OR keyword)",
        "AI offers (ESCO AND keyword)",
        "% AI of all offers",
        "% AI of ICT offers",
        "",
        "Strict/Core AI offers",
        "Extended AI only offers",
    ],
    "Value": [
        total_offers,
        n_ict,
        f"{n_ict / total_offers * 100:.2f}%",
        "",
        int(df["has_ai_esco"].sum()),
        int(df["has_ai_kw_strict"].sum()),
        int(df["is_ai_offer"].sum()),
        int((df["has_ai_esco"] & df["has_ai_kw_strict"]).sum()),
        f"{df['is_ai_offer'].mean() * 100:.2f}%",
        f"{n_ai_ict / n_ict * 100:.2f}%" if n_ict > 0 else "N/A",
        "",
        int(df["is_ai_strict"].sum()),
        int(df["is_ai_extended_only"].sum()),
    ],
}
overview_df = pd.DataFrame(overview_data)

# ── 6.2 AI skills frequency ─────────────────────────────────────────────────
skill_counts = Counter()
skill_counts_ict = Counter()
ict_ai_offers = df[df["is_ai_and_ict"]]
for skills in ai_offers["ai_skills_esco"]:
    for s in skills:
        key = (s["esco_en"], s["esco_pl"], s["category"], s["scope"])
        skill_counts[key] += 1
for skills in ict_ai_offers["ai_skills_esco"]:
    for s in skills:
        key = (s["esco_en"], s["esco_pl"], s["category"], s["scope"])
        skill_counts_ict[key] += 1

skills_freq = pd.DataFrame(
    [
        {
            "ESCO Skill (EN)": k[0],
            "ESCO Skill (PL)": k[1],
            "AI Category": k[2],
            "Scope": k[3],
            "N offers": v,
            "% of AI offers": v / len(ai_offers) * 100,
            "% of ICT offers": skill_counts_ict.get(k, 0) / n_ict * 100 if n_ict > 0 else 0,
            "% of all offers": v / total_offers * 100,
        }
        for k, v in skill_counts.most_common()
    ]
)

# Keyword frequency
kw_counts = Counter()
for kws in ai_offers["ai_keywords_found"]:
    for kw in kws:
        kw_counts[kw.lower()] += 1

kw_freq = pd.DataFrame(
    [
        {
            "Keyword": k,
            "N offers": v,
            "% of AI offers": v / len(ai_offers) * 100,
        }
        for k, v in kw_counts.most_common()
    ]
)

# ── 6.3 Co-occurring ESCO skills ─────────────────────────────────────────────
print("Analyzing co-occurring skills...")

ai_all_skills = Counter()
non_ai_all_skills = Counter()

for _, row in ai_offers.iterrows():
    seen = set()
    for s in row["skills_parsed"]:
        label = s.get("esco", "").strip().lower()
        if label and label not in seen:
            seen.add(label)
            ai_all_skills[label] += 1

sample_non_ai = non_ai.sample(n=min(100_000, len(non_ai)), random_state=42)
for _, row in sample_non_ai.iterrows():
    seen = set()
    for s in row["skills_parsed"]:
        label = s.get("esco", "").strip().lower()
        if label and label not in seen:
            seen.add(label)
            non_ai_all_skills[label] += 1

n_ai = len(ai_offers)
n_non_ai = len(sample_non_ai)

cooccur_data = []
for skill, ai_count in ai_all_skills.items():
    if skill in ai_skills_pl:
        continue
    non_ai_count = non_ai_all_skills.get(skill, 0)
    ai_rate = ai_count / n_ai
    non_ai_rate = non_ai_count / n_non_ai if n_non_ai > 0 else 0
    ratio = ai_rate / non_ai_rate if non_ai_rate > 0 else float("inf")
    is_transversal = skill in transversal_set
    en_label = skill_pl_to_en.get(skill, "")
    cooccur_data.append(
        {
            "ESCO Skill (PL)": skill,
            "ESCO Skill (EN)": en_label,
            "Transversal": "Yes" if is_transversal else "No",
            "N in AI offers": ai_count,
            "% in AI offers": ai_rate * 100,
            "% in non-AI offers": non_ai_rate * 100,
            "Overrepresentation ratio": round(ratio, 2),
        }
    )

cooccur_df = pd.DataFrame(cooccur_data)
cooccur_df = cooccur_df[cooccur_df["N in AI offers"] >= 50].sort_values(
    "Overrepresentation ratio", ascending=False
)

cooccur_transversal = cooccur_df[cooccur_df["Transversal"] == "Yes"].copy()
cooccur_other = cooccur_df[cooccur_df["Transversal"] == "No"].head(60).copy()

# ── 6.4 Technologies ─────────────────────────────────────────────────────────
print("Analyzing technologies...")


def parse_tech(val):
    if not val or pd.isna(val):
        return []
    return [t.strip() for t in str(val).split(",") if t.strip()]


ai_tech = Counter()
non_ai_tech = Counter()

for techs in ai_offers["technologies"].apply(parse_tech):
    for t in techs:
        ai_tech[t.lower()] += 1

for techs in sample_non_ai["technologies"].apply(parse_tech):
    for t in techs:
        non_ai_tech[t.lower()] += 1

tech_data = []
for tech, ai_count in ai_tech.items():
    if ai_count < 20:
        continue
    non_ai_count = non_ai_tech.get(tech, 0)
    ai_rate = ai_count / n_ai
    non_ai_rate = non_ai_count / n_non_ai if n_non_ai > 0 else 0
    ratio = ai_rate / non_ai_rate if non_ai_rate > 0 else float("inf")
    tech_data.append(
        {
            "Technology": tech,
            "N in AI offers": ai_count,
            "% in AI offers": ai_rate * 100,
            "% in non-AI offers": non_ai_rate * 100,
            "Overrepresentation ratio": round(ratio, 2),
        }
    )

tech_df = pd.DataFrame(tech_data).sort_values("Overrepresentation ratio", ascending=False)

# ── 6.5 Contract types ──────────────────────────────────────────────────────
print("Analyzing contracts...")

CONTRACT_MAP = {
    "umowa o pracę": "Employment contract (UoP)",
    "contract of employment": "Employment contract (UoP)",
    "umowa o pracę tymczasową": "Temporary employment",
    "umowa na zastępstwo": "Temporary employment",
    "kontrakt B2B": "B2B contract",
    "b2b contract": "B2B contract",
    "umowa zlecenie": "Contract of mandate (zlecenie)",
    "contract of mandate": "Contract of mandate (zlecenie)",
    "umowa o dzieło": "Contract for specific work (o dzieło)",
    "umowa agencyjna": "Agency contract",
    "umowa o staż / praktyki": "Internship / traineeship",
}


def map_contracts(val):
    if not val or pd.isna(val):
        return ["No data"]
    raw = [c.strip() for c in str(val).split(",")]
    mapped = set()
    for c in raw:
        mapped.add(CONTRACT_MAP.get(c.strip(), c.strip()))
    return list(mapped)


ai_contracts_agg = Counter()
for contracts in ai_offers["contract_types"].apply(map_contracts):
    for c in contracts:
        ai_contracts_agg[c] += 1

all_contracts_agg = Counter()
for contracts in df["contract_types"].apply(map_contracts):
    for c in contracts:
        all_contracts_agg[c] += 1

contract_data = []
for c, ai_cnt in ai_contracts_agg.most_common():
    all_cnt = all_contracts_agg.get(c, 0)
    contract_data.append(
        {
            "Contract Type": c,
            "N in AI offers": ai_cnt,
            "% in AI offers": ai_cnt / len(ai_offers) * 100,
            "% in all offers": all_cnt / total_offers * 100,
        }
    )
contract_df = pd.DataFrame(contract_data)

# ── 6.6 Seniority ───────────────────────────────────────────────────────────
print("Analyzing seniority...")

SENIORITY_MAP = {
    "junior": [
        "młodszy specjalista (junior)", "junior specialist (junior)",
        "praktykant / stażysta", "trainee", "asystent", "assistant",
        "entry level & blue collar",
    ],
    "mid": [
        "specjalista (mid / regular)", "specialist (mid / regular)",
    ],
    "senior": [
        "starszy specjalista (senior)", "senior specialist (senior)",
        "ekspert", "expert",
    ],
    "management": [
        "kierownik / koordynator", "manager / supervisor",
        "menedżer", "team manager",
        "dyrektor", "director", "prezes", "executive",
    ],
    "blue collar": [
        "pracownik fizyczny",
    ],
}

seniority_lookup = {}
for group, labels in SENIORITY_MAP.items():
    for l in labels:
        seniority_lookup[l] = group


def map_seniority(val):
    if not val or pd.isna(val):
        return ["No data"]
    parts = [p.strip().lower() for p in str(val).split(",")]
    mapped = set()
    for p in parts:
        mapped.add(seniority_lookup.get(p, "Other"))
    return list(mapped)


ai_seniority_agg = Counter()
for levels in ai_offers["seniority_level"].apply(map_seniority):
    for l in levels:
        ai_seniority_agg[l] += 1

all_seniority_agg = Counter()
for levels in df["seniority_level"].apply(map_seniority):
    for l in levels:
        all_seniority_agg[l] += 1

SENIORITY_ORDER = ["junior", "mid", "senior", "management", "blue collar", "No data", "Other"]
seniority_data = []
for level in SENIORITY_ORDER:
    ai_cnt = ai_seniority_agg.get(level, 0)
    all_cnt = all_seniority_agg.get(level, 0)
    if ai_cnt == 0 and all_cnt == 0:
        continue
    seniority_data.append(
        {
            "Seniority Level": level.title(),
            "N in AI offers": ai_cnt,
            "% in AI offers": ai_cnt / len(ai_offers) * 100,
            "% in all offers": all_cnt / total_offers * 100,
        }
    )
seniority_df = pd.DataFrame(seniority_data)

# ── 6.7 Location ────────────────────────────────────────────────────────────
print("Analyzing locations...")

CITY_PATTERNS = {
    "Warszawa": r"warszaw",
    "Kraków": r"krak[oó]w",
    "Wrocław": r"wroc[lł]aw",
    "Poznań": r"pozna[nń]",
    "Gdańsk": r"gda[nń]sk",
    "Łódź": r"[lł][oó]d[zź]",
    "Katowice": r"katowic",
    "Lublin": r"lublin",
    "Szczecin": r"szczecin",
    "Bydgoszcz": r"bydgoszcz",
    "Białystok": r"bia[lł]ystok",
    "Gdynia": r"gdyni",
    "Rzeszów": r"rzesz[oó]w",
    "Toruń": r"toru[nń]",
    "Kielce": r"kielc",
    "Opole": r"opol",
    "Olsztyn": r"olsztyn",
    "Zielona Góra": r"zielona g[oó]r",
    "Gliwice": r"gliwic",
    "Częstochowa": r"cz[eę]stochow",
    "Radom": r"\bradom",
    "Sosnowiec": r"sosnowiec",
    "Bielsko-Biała": r"bielsko",
    "Praca zdalna": r"(zdaln|remote|home office)",
    "Zagranica": r"zagrani",
}


def extract_city(location):
    if not location or pd.isna(location):
        return "brak danych"
    loc_lower = str(location).lower()
    cities_found = []
    for city, pattern in CITY_PATTERNS.items():
        if re.search(pattern, loc_lower):
            cities_found.append(city)
    if cities_found:
        return "; ".join(cities_found)
    return "inne"


ai_offers_loc = ai_offers.copy()
ai_offers_loc["city"] = ai_offers_loc["location"].apply(extract_city)

city_counts = Counter()
for cities_str in ai_offers_loc["city"]:
    for c in cities_str.split("; "):
        city_counts[c.strip()] += 1

all_loc = df.copy()
all_loc["city"] = all_loc["location"].apply(extract_city)
all_city_counts = Counter()
for cities_str in all_loc["city"]:
    for c in cities_str.split("; "):
        all_city_counts[c.strip()] += 1

location_data = []
for city, cnt in city_counts.most_common():
    all_cnt = all_city_counts.get(city, 0)
    location_data.append(
        {
            "City": city,
            "N in AI offers": cnt,
            "% in AI offers": cnt / len(ai_offers) * 100,
            "% in all offers": all_cnt / total_offers * 100,
        }
    )
location_df = pd.DataFrame(location_data)

# ── 6.8 ESCO job titles & NACE ──────────────────────────────────────────────
print("Analyzing job titles and NACE codes...")

ai_ids = set(ai_offers["id"].values)
nace_ai = nace_df[nace_df["job_id"].isin(ai_ids)]

esco_title_counts = nace_ai["esco_label"].value_counts().head(50).reset_index()
esco_title_counts.columns = ["ESCO Job Title (PL)", "N offers"]
esco_title_counts["ESCO Job Title (EN)"] = esco_title_counts["ESCO Job Title (PL)"].apply(
    lambda x: occ_pl_to_en.get(str(x).strip().lower(), "")
)
esco_title_counts["% of AI offers"] = esco_title_counts["N offers"] / len(ai_offers) * 100
esco_title_counts = esco_title_counts[
    ["ESCO Job Title (PL)", "ESCO Job Title (EN)", "N offers", "% of AI offers"]
]

nace_counts = (
    nace_ai.groupby(["nace_code", "nace_title"])
    .size()
    .reset_index(name="N offers")
    .sort_values("N offers", ascending=False)
    .head(50)
)
nace_counts["% of AI offers"] = nace_counts["N offers"] / len(ai_offers) * 100

# NACE section level (first letter)
nace_ai_section = nace_ai.copy()
nace_ai_section["nace_section"] = nace_ai_section["nace_code"].str[0]

NACE_SECTIONS = {
    "A": "Agriculture, forestry and fishing",
    "B": "Mining and quarrying",
    "C": "Manufacturing",
    "D": "Electricity, gas",
    "E": "Water supply, sewerage",
    "F": "Construction",
    "G": "Wholesale and retail trade",
    "H": "Transportation and storage",
    "I": "Accommodation and food service",
    "J": "Information and communication",
    "K": "Financial and insurance",
    "L": "Real estate",
    "M": "Professional, scientific and technical",
    "N": "Administrative and support service",
    "O": "Public administration and defence",
    "P": "Education",
    "Q": "Human health and social work",
    "R": "Arts, entertainment and recreation",
    "S": "Other service activities",
}

nace_section_counts = nace_ai_section["nace_section"].value_counts().reset_index()
nace_section_counts.columns = ["NACE Section", "N offers"]
nace_section_counts["Section Name"] = nace_section_counts["NACE Section"].map(NACE_SECTIONS)
nace_section_counts["% of AI offers"] = nace_section_counts["N offers"] / len(ai_offers) * 100
nace_section_counts = nace_section_counts[
    ["NACE Section", "Section Name", "N offers", "% of AI offers"]
].sort_values("N offers", ascending=False)

# ── 6.9 AI as % of ICT breakdown ─────────────────────────────────────────────
print("Analyzing AI within ICT...")

ict_offers = df[df["is_ict"]]

ict_breakdown_data = {
    "Metric": [
        "Total ICT offers",
        "AI offers within ICT",
        "% AI within ICT",
        "Core/Strict AI within ICT",
        "Extended AI only within ICT",
        "",
        "AI offers outside ICT",
        "Total AI offers",
    ],
    "Value": [
        n_ict,
        n_ai_ict,
        f"{n_ai_ict / n_ict * 100:.2f}%" if n_ict > 0 else "N/A",
        int((df["is_ai_strict"] & df["is_ict"]).sum()),
        int((df["is_ai_extended_only"] & df["is_ict"]).sum()),
        "",
        int((df["is_ai_offer"] & ~df["is_ict"]).sum()),
        int(df["is_ai_offer"].sum()),
    ],
}
ict_breakdown_df = pd.DataFrame(ict_breakdown_data)

# AI skill frequency as % of ICT
ict_skill_freq = pd.DataFrame(
    [
        {
            "ESCO Skill (EN)": k[0],
            "ESCO Skill (PL)": k[1],
            "AI Category": k[2],
            "Scope": k[3],
            "N in ICT offers": skill_counts_ict.get(k, 0),
            "% of ICT offers": skill_counts_ict.get(k, 0) / n_ict * 100 if n_ict > 0 else 0,
        }
        for k, v in skill_counts.most_common()
        if skill_counts_ict.get(k, 0) > 0
    ]
)

# Keyword frequency within ICT
kw_counts_ict = Counter()
for kws in ict_ai_offers["ai_keywords_found"]:
    for kw in kws:
        kw_counts_ict[kw.lower()] += 1

ict_kw_freq = pd.DataFrame(
    [
        {
            "Keyword": k,
            "N in ICT offers": v,
            "% of ICT offers": v / n_ict * 100 if n_ict > 0 else 0,
        }
        for k, v in kw_counts_ict.most_common()
    ]
)

# ── 7. EXPORT TO EXCEL ──────────────────────────────────────────────────────
print("\nExporting to Excel...")

output_path = "ai_demand_analysis.xlsx"

SHEETS = {
    "1. Overview": overview_df,
    "2. AI Skills (ESCO)": skills_freq,
    "3. AI Keywords": kw_freq,
    "4. Co-occurring Skills": cooccur_other,
    "5. Co-occur Transversal": cooccur_transversal,
    "6. Technologies": tech_df,
    "7. Contract Type": contract_df,
    "8. Seniority": seniority_df,
    "9. Location": location_df,
    "10. ESCO Job Titles": esco_title_counts,
    "11. NACE Codes": nace_counts,
    "12. NACE Sections": nace_section_counts,
    "13. AI within ICT": ict_breakdown_df,
    "14. AI Skills (% ICT)": ict_skill_freq,
    "15. AI Keywords (% ICT)": ict_kw_freq,
}

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    for sheet_name, data in SHEETS.items():
        data.to_excel(writer, sheet_name=sheet_name, index=False)

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if isinstance(cell.value, float):
                    cell.number_format = "#,##0.00"
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row in ws.iter_rows(
                min_row=1, max_row=min(ws.max_row, 100), min_col=col_idx, max_col=col_idx
            ):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 55)

print(f"\nDone! Saved to: {output_path}")
print(f"\nKey stats:")
print(f"  Total offers: {total_offers:,}")
print(f"  AI offers (combined): {df['is_ai_offer'].sum():,} ({df['is_ai_offer'].mean()*100:.2f}%)")
print(f"  - ESCO match: {df['has_ai_esco'].sum():,}")
print(f"  - Keyword match: {df['has_ai_kw_strict'].sum():,}")
print(f"  - Both: {(df['has_ai_esco'] & df['has_ai_kw_strict']).sum():,}")
