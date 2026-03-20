import pandas as pd
import re

df = pd.read_csv("ESCO dataset - v1.2.1 - classification - en - csv/skills_en.csv")

text_cols = ["preferredLabel", "altLabels", "description", "definition", "scopeNote"]
for col in text_cols:
    df[col] = df[col].fillna("")
df["all_text"] = df[text_cols].apply(lambda x: " ".join(x).lower(), axis=1)

ai_keywords = [
    "artificial intelligence",
    "machine learning",
    "deep learning",
    "neural network",
    "natural language processing",
    "computer vision",
    "data science",
    "data mining",
    "big data",
    "predictive model",
    "predictive analytics",
    "reinforcement learning",
    "supervised learning",
    "unsupervised learning",
    "generative ai",
    "generative model",
    "chatbot",
    "speech recognition",
    "image recognition",
    "pattern recognition",
    "recommendation system",
    "recommender system",
    "cognitive computing",
    "tensorflow",
    "pytorch",
    "scikit-learn",
    "keras",
    "language model",
    "large language model",
    "transformer model",
    "sentiment analysis",
    "text mining",
    "information extraction",
    "knowledge graph",
    "expert system",
    "genetic algorithm",
    "bayesian",
    "convolutional neural",
    "recurrent neural",
    "transfer learning",
    "federated learning",
    "autonomous system",
    "autonomous vehicle",
    "self-driving",
    "facial recognition",
    "biometric",
    "data annotation",
    "training data",
    "machine translation",
    "dialogue system",
    "intelligent system",
    "intelligent agent",
    "robotics",
    "robotic process automation",
    "internet of things",
    "edge computing",
    "cloud computing",
    "statistical model",
    "statistical learning",
    "feature engineering",
    "model training",
    "classification model",
    "regression model",
    "clustering algorithm",
]

matches = set()
for kw in ai_keywords:
    mask = df["all_text"].str.contains(re.escape(kw.lower()), case=False, na=False)
    matches.update(df[mask].index.tolist())

ai_df = df.loc[sorted(list(matches))].copy()

CATEGORY_RULES = {
    "Core AI & Machine Learning": [
        "principles of artificial intelligence",
        "machine learning",
        "deep learning",
        "artificial neural networks",
        "cognitive computing",
        "utilise machine learning",
        "perform dimensionality reduction",
        "emergent technologies",
    ],
    "Natural Language Processing": [
        "natural language processing",
        "machine translation",
        "information extraction",
        "speech recognition",
        "improve translated texts",
    ],
    "Computer Vision & Image Recognition": [
        "computer vision",
        "image recognition",
        "develop computer vision system",
    ],
    "Data Science & Analytics": [
        "data science",
        "data mining",
        "data mining methods",
        "perform data mining",
        "analyse big data",
        "analyse large-scale data in healthcare",
        "statistics",
        "statistical analysis system software",
        "statistical modeling techniques",
        "apply statistical analysis techniques",
        "use methods of logistical data analysis",
        "business analytics",
        "unstructured data",
        "build predictive models",
        "develop predictive models",
    ],
    "Robotics & Autonomous Systems": [
        "robotics",
        "human-robot collaboration",
    ],
    "IoT, Cloud & Edge Computing": [
        "internet of things",
        "cloud technologies",
        "cloud monitoring and reporting",
        "digital twin technology",
        "smart city features",
    ],
    "Biometrics & Security": [
        "biometrics",
        "contribute to development of biometric systems",
        "maltego",
    ],
    "AI Applications & Domain-Specific": [
        "predictive maintenance",
        "build recommender systems",
        "guide learners in using assistive technologies",
        "teach computer science",
    ],
}


def assign_category(label):
    label_lower = label.lower().strip()
    for category, labels in CATEGORY_RULES.items():
        if label_lower in [l.lower() for l in labels]:
            return category
    return "Other / Uncategorized"


ai_df["ai_category"] = ai_df["preferredLabel"].apply(assign_category)

uncategorized = ai_df[ai_df["ai_category"] == "Other / Uncategorized"]
if len(uncategorized) > 0:
    print("WARNING - Uncategorized skills:")
    for _, row in uncategorized.iterrows():
        print(f"  {row['preferredLabel']}")
    print()

output = ai_df[
    [
        "preferredLabel",
        "skillType",
        "ai_category",
        "conceptUri",
        "reuseLevel",
        "description",
        "altLabels",
    ]
].copy()

output.columns = [
    "ESCO Skill/Knowledge Name",
    "ESCO Type",
    "AI Category",
    "ESCO URI",
    "Reuse Level",
    "Description",
    "Alternative Labels",
]

output = output.sort_values(["AI Category", "ESCO Type", "ESCO Skill/Knowledge Name"])

output_path = "ai_skills_esco_classification.xlsx"
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    output.to_excel(writer, sheet_name="AI Skills Classification", index=False)

    summary = (
        output.groupby("AI Category")
        .agg(
            Total=("ESCO Skill/Knowledge Name", "count"),
            Knowledge=("ESCO Type", lambda x: (x == "knowledge").sum()),
            Skills_Competences=("ESCO Type", lambda x: (x == "skill/competence").sum()),
        )
        .reset_index()
    )
    summary.loc[len(summary)] = [
        "TOTAL",
        summary["Total"].sum(),
        summary["Knowledge"].sum(),
        summary["Skills_Competences"].sum(),
    ]
    summary.to_excel(writer, sheet_name="Summary by Category", index=False)

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            adjusted = min(max_length + 2, 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    ws_main = writer.sheets["AI Skills Classification"]
    category_fill = {
        "Core AI & Machine Learning": PatternFill(start_color="D6E4F0", fill_type="solid"),
        "Natural Language Processing": PatternFill(start_color="E2EFDA", fill_type="solid"),
        "Computer Vision & Image Recognition": PatternFill(start_color="FCE4D6", fill_type="solid"),
        "Data Science & Analytics": PatternFill(start_color="EDEDED", fill_type="solid"),
        "Robotics & Autonomous Systems": PatternFill(start_color="FFF2CC", fill_type="solid"),
        "IoT, Cloud & Edge Computing": PatternFill(start_color="D9E2F3", fill_type="solid"),
        "Biometrics & Security": PatternFill(start_color="F2DCDB", fill_type="solid"),
        "AI Applications & Domain-Specific": PatternFill(start_color="E2D9F3", fill_type="solid"),
    }
    for row in ws_main.iter_rows(min_row=2, max_row=ws_main.max_row, max_col=ws_main.max_column):
        cat = row[2].value
        if cat in category_fill:
            for cell in row:
                cell.fill = category_fill[cat]

print(f"Total AI-related skills: {len(output)}")
print()
print("Summary by category:")
print(summary.to_string(index=False))
print(f"\nSaved to: {output_path}")
