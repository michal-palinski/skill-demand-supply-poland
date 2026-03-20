import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── DATA ──────────────────────────────────────────────────────────────────────
# URL1: https://www.pracuj.pl/praca?lang=uk
# URL2: https://www.pracuj.pl/praca?ua=true&lang=uk  (from screenshots)

data = {
    "URL1 – wszyscy (lang=uk)": {
        "Poziom stanowiska": [
            ("praktykant / praktykantka - stażysta / stażystka", 626),
            ("asystent / asystentka", 2207),
            ("młodszy specjalista / młodsza specjalistka (junior)", 6953),
            ("specjalista / specjalistka (mid / regular)", 31066),
            ("starszy specjalista / starsza specjalistka (senior)", 8653),
            ("ekspert / ekspertka", 1960),
            ("kierownik / kierowniczka - koordynator / koordynatorka", 6169),
            ("menedżer / menedżerka", 3252),
            ("dyrektor / dyrektorka", 810),
            ("prezes / prezeska", 25),
            ("pracownik fizyczny / pracowniczka fizyczna", 13075),
        ],
        "Rodzaj umowy": [
            ("umowa o pracę", 50187),
            ("umowa o dzieło", 623),
            ("umowa zlecenie", 10891),
            ("kontrakt B2B", 17980),
            ("umowa na zastępstwo", 558),
            ("umowa agencyjna", 810),
            ("umowa o pracę tymczasową", 500),
            ("umowa o staż / praktyki", 404),
        ],
        "Wymiar pracy": [
            ("część etatu", 5710),
            ("dodatkowa / tymczasowa", 1917),
            ("pełny etat", 59869),
        ],
        "Tryb pracy": [
            ("praca stacjonarna", 41268),
            ("praca hybrydowa", 17244),
            ("praca zdalna", 5899),
            ("praca mobilna", 9096),
        ],
        "Wynagrodzenie": [
            ("Pokaż oferty z widełkami wynagrodzeń", 21124),
        ],
    },
    "URL2 – Ukraińcy (ua=true)": {
        "Poziom stanowiska": [
            ("praktykant / praktykantka - stażysta / stażystka", 48),
            ("asystent / asystentka", 121),
            ("młodszy specjalista / młodsza specjalistka (junior)", 429),
            ("specjalista / specjalistka (mid / regular)", 1337),
            ("starszy specjalista / starsza specjalistka (senior)", 529),
            ("ekspert / ekspertka", 98),
            ("kierownik / kierowniczka - koordynator / koordynatorka", 422),
            ("menedżer / menedżerka", 93),
            ("dyrektor / dyrektorka", 12),
            ("prezes / prezeska", 0),
            ("pracownik fizyczny / pracowniczka fizyczna", 2516),
        ],
        "Rodzaj umowy": [
            ("umowa o pracę", 3301),
            ("umowa o dzieło", 53),
            ("umowa zlecenie", 1545),
            ("kontrakt B2B", 1214),
            ("umowa na zastępstwo", 17),
            ("umowa agencyjna", 47),
            ("umowa o pracę tymczasową", 113),
            ("umowa o staż / praktyki", 67),
        ],
        "Wymiar pracy": [
            ("część etatu", 874),
            ("dodatkowa / tymczasowa", 339),
            ("pełny etat", 4628),
        ],
        "Tryb pracy": [
            ("praca stacjonarna", 3325),
            ("praca hybrydowa", 992),
            ("praca zdalna", 359),
            ("praca mobilna", 684),
        ],
        "Wynagrodzenie": [
            ("Pokaż oferty z widełkami wynagrodzeń", 2653),
        ],
    },
}

# ── BUILD DATAFRAMES ──────────────────────────────────────────────────────────
def build_df(sections: dict) -> pd.DataFrame:
    rows = []
    for section, items in sections.items():
        total = sum(count for _, count in items)
        for label, count in items:
            pct = round(count / total * 100, 1) if total > 0 else 0.0
            rows.append({
                "Sekcja": section,
                "Opcja": label,
                "Liczba ofert": count,
                "% w sekcji": pct,
            })
        # Section subtotal row
        rows.append({
            "Sekcja": section,
            "Opcja": "SUMA",
            "Liczba ofert": total,
            "% w sekcji": 100.0,
        })
    return pd.DataFrame(rows)

output_path = "pracuj_filters_final.xlsx"

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    for sheet_name, sections in data.items():
        df = build_df(sections)
        safe_name = sheet_name[:31]
        df.to_excel(writer, sheet_name=safe_name, index=False)

# ── STYLING ───────────────────────────────────────────────────────────────────
wb = load_workbook(output_path)

HEADER_FILL  = PatternFill("solid", fgColor="3B4A6B")
SECTION_FILL = PatternFill("solid", fgColor="D6DCF0")
TOTAL_FILL   = PatternFill("solid", fgColor="B8C2E8")
ALT_FILL     = PatternFill("solid", fgColor="F5F6FB")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(bold=True, color="1A2347", size=10)
TOTAL_FONT   = Font(bold=True, italic=True, color="1A2347", size=10)
BODY_FONT    = Font(color="2C2C2C", size=10)

thin = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_sheet(ws):
    # Column widths
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 56
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14

    # Header row
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22

    # Body rows
    section_color_toggle = {}
    color_idx = 0
    section_fills = {}

    for row in ws.iter_rows(min_row=2):
        section = row[0].value
        if section not in section_fills:
            section_fills[section] = color_idx % 2
            color_idx += 1

        is_total = row[1].value == "SUMA"

        for cell in row:
            cell.border = THIN_BORDER
            if is_total:
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT
                cell.alignment = Alignment(horizontal="center" if cell.column >= 3 else "left", vertical="center")
            elif cell.column == 1:
                cell.fill = SECTION_FILL
                cell.font = SECTION_FONT
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                fill = ALT_FILL if section_fills[section] else WHITE_FILL
                cell.fill = fill
                cell.font = BODY_FONT
                cell.alignment = Alignment(
                    horizontal="center" if cell.column >= 3 else "left",
                    vertical="center",
                    wrap_text=True,
                )

        # Format % column
        pct_cell = row[3]
        if pct_cell.value is not None and not is_total:
            pct_cell.number_format = '0.0"%"'

    ws.freeze_panes = "A2"

for sheet_name in wb.sheetnames:
    style_sheet(wb[sheet_name])

wb.save(output_path)
print(f"Saved: {output_path}")

# Print summary
for sheet_name, sections in data.items():
    print(f"\n{'='*60}")
    print(f"  {sheet_name}")
    print(f"{'='*60}")
    for section, items in sections.items():
        total = sum(c for _, c in items)
        print(f"\n  {section}  (suma: {total})")
        for label, count in items:
            pct = count / total * 100 if total else 0
            print(f"    {label:<55} {count:>5}  ({pct:.1f}%)")
